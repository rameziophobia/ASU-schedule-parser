# written by Ramez
# https://github.com/rameziophobia 
import openpyxl 

requested_courses = ['CSE365', 'CSE335', 'CSE325', 'CSE316', 'CSE326', 'CSE415', 'CSE436', 'CSE386', "CSE320"]
data = ""
with open('exams.txt', 'r') as file:
    data = file.read().replace('\n', ' ')

pages = data.split('AIN SHAMS UNIVERSITY FACULTY OF ENGINEERING CREDIT HOURS ENGINEERING PROGRAMS Midterm EXAMINATION Fall 2020 DAY AND DATE CODE COURSE STUD Time HALL')

class Exam():
    
    def __init__(self, pages):
        self.pages = pages
        self.exams = {}
        self.process_pages()

    def process_pages(self):
        for page in self.pages:
            self.process_page(page)

    def process_page(self, page):
        if len(page) != 0:
            page_string = page.strip().split(' ')
            if page_string[1] == "15-11-2019":
                page_string[1] = "4-12-2020"
            day = page_string[:2]
            exam_strings = []
            for word in page_string:
                is_a_subject_code, subject_code = self.get_code_if_valid(word)
                if is_a_subject_code:
                    self.add_exam(exam_strings)
                    exam_strings = [*day, subject_code]
                else:
                    exam_strings.append(word)

            self.add_exam(exam_strings)

    def add_exam(self, exam_strings):
        if len(exam_strings) >= 3:
            try:
                hours_index = exam_strings.index('–')
            except ValueError:
                hours_index = -1
            
            if hours_index != -1:
                self.last_exam_time = ''.join([*exam_strings[hours_index - 1: hours_index + 2]])
                self.last_exam_hall = ' '.join([*exam_strings[hours_index + 2:]])

            exam_strings = [*exam_strings[:3], 
                ' '.join([*exam_strings[3:hours_index - 2]]), 
                self.last_exam_time,
                self.last_exam_hall,
                ]

            if exam_strings[2] in self.exams:
                print("code conflict for subject " + exam_strings[2])
            self.exams[exam_strings[2]] = exam_strings

    def get_code_if_valid(self, word):
        if len(word) == 8 or len(word) == 6:
            subject_code_letters = word[:3]
            subject_code_numbers = word[len(word) - 3:]
            if subject_code_letters.isupper() and subject_code_letters.isalpha() and subject_code_numbers.isnumeric():
                return True, word
        return False, ''

    def get_exam_schedule(self, codes):
        exams = [self.exams[code] for code in codes]
        exams.sort(key=lambda e: self.get_date_for_sort(e[1]))
        return exams

    def get_date_for_sort(self, date_string):
        day, month, year = date_string.split("-")
        return year + month + day

exam = Exam(pages)
results = exam.get_exam_schedule(requested_courses)
[print(ex) for ex in results]

wb = openpyxl.load_workbook("temp1.xlsx")
ws = wb.active
for index, row in enumerate(results):
    row_num = str(index + 4)
    ws['B' + row_num] = row[0]
    ws['C' + row_num] = row[1]
    ws['D' + row_num] = row[2] + " " + row[3]
    ws['E' + row_num] = row[4]
    ws['F' + row_num] = row[5]

wb.save('temp2.xlsx')